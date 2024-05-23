import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import pygame
import os

class Login_System:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title('Welcome to Sigma University')
        self.root.geometry('925x500+300+200')
        self.root.configure(bg="#FFFFFF")
        self.root.resizable(False, False)

        # Play Music
        pygame.init()
        pygame.mixer.init()
        music = r"C:\Users\hatab\Downloads\getmogged.mp3"
        if os.path.exists(music):
            try:
                pygame.mixer.music.load(music)
                # Loop Song
                pygame.mixer.music.play(-1)
            except pygame.error as e:
                messagebox.showwarning("Warning", f"Failed to load music file: {e}")
        else:
            messagebox.showwarning("Warning", "Music file not found. Music will not play.")

        self.show_login_screen()
        self.root.mainloop()

    def show_login_screen(self):
        self.clear_window()

        # Logo Image
        img_path = r"C:\Users\hatab\OneDrive\Pictures\Saved Pictures\Sigma-sigma.jpg"
        if os.path.exists(img_path):
            imgSigma = Image.open(img_path)
            self.imgSigma = ImageTk.PhotoImage(imgSigma)
            tk.Label(self.root, image=self.imgSigma, bg="white").place(x=50, y=150)
        else:
            messagebox.showwarning("Warning", "Logo image not found. Image will not be displayed.")

        # Place to put Header and User Input
        self.login_frame = tk.Frame(self.root, width=350, height=350, bg="white")
        self.login_frame.place(x=480, y=70)

        # Header GUI
        heading = tk.Label(self.login_frame, text="Sign in", fg="#57a1f8", bg="white", font=("Centaur", 23, "bold"))
        heading.place(x=125, y=5)

        # Input Username
        username = tk.Entry(self.login_frame, width=25, fg="black", border=0, bg="white", font=("Jokerman", 11))
        username.place(x=30, y=80)
        username.insert(0, "Username")

        # Interact When Click Username
        def when_enter(e):
            username.delete(0, "end")

        def when_leave(e):
            name = username.get()
            if name == "":
                username.insert(0, "Username")

        username.bind("<FocusIn>", when_enter)
        username.bind("<FocusOut>", when_leave)

        # Underline Username
        tk.Frame(self.login_frame, width=295, height=2, bg="black").place(x=25, y=107)

        # Input Password
        password = tk.Entry(self.login_frame, width=25, fg="black", border=0, bg="white", font=("Jokerman", 11))
        password.place(x=30, y=150)
        password.insert(0, "Password")

        # Interact Password When Click
        def when_enter(e):
            password.delete(0, "end")
            password.config(show="*")

        def when_leave(e):
            passcode = password.get()
            if passcode == "":
                password.insert(0, "Password")
                password.config(show="")

        password.bind("<FocusIn>", when_enter)
        password.bind("<FocusOut>", when_leave)

        # Underline Password
        tk.Frame(self.login_frame, width=295, height=2, bg="black").place(x=25, y=177)

        # Sign In
        sign_in = tk.Button(self.login_frame, width=39, pady=7, text="Sign In", bg="#57a1f8", fg="white", command=lambda: self.login(username, password), border=0)
        sign_in.place(x=35, y=204)

        # Label "No Account"
        no_account = tk.Label(self.login_frame, text="No account?", bg="white", fg="black", font=("Centaur", 10))
        no_account.place(x=200, y=250)

        sign_up = tk.Button(self.login_frame, width=6, text="Sign Up", command=self.show_register_screen, bg="White", fg="#57a1f8", cursor="hand2", border=0)
        sign_up.place(x=270, y=250)

        # Volume Control Slider
        volume_label = tk.Label(self.root, text="Volume", bg="white", fg="black", font=("Centaur", 10))
        volume_label.place(x=30, y=425)

        self.volume_slider = tk.Scale(self.root, from_=0, to_=1, orient='horizontal', resolution=0.1, bg="white", command=self.set_volume)
        
        # Volume Max
        self.volume_slider.set(1)
        self.volume_slider.place(x=5, y=445)

    def set_volume(self, val):
        volume = float(val)
        pygame.mixer.music.set_volume(volume)

    def show_register_screen(self):
        self.clear_window()

        # Image
        img_path = r"C:\Users\hatab\OneDrive\Pictures\Saved Pictures\Top G.png"
        if os.path.exists(img_path):
            imgTopG = Image.open(img_path)
            self.imgTopG = ImageTk.PhotoImage(imgTopG)
            tk.Label(self.root, image=self.imgTopG, bg="white").place(x=50, y=80)
        else:
            messagebox.showwarning("Warning", "Logo image not found. Image will not be displayed.")

        # Place to put Header and User Input
        self.register_frame = tk.Frame(self.root, width=350, height=350, bg="white")
        self.register_frame.place(x=480, y=70)

        # Header GUI
        heading = tk.Label(self.register_frame, text="Sign Up", fg="#57a1f8", bg="white", font=("Centaur", 23, "bold"))
        heading.place(x=120, y=5)

        # Input Username
        username = tk.Entry(self.register_frame, width=25, fg="black", border=0, bg="white", font=("Jokerman", 11))
        username.place(x=30, y=80)
        username.insert(0, "Username")

        # Interact When Click Username
        def when_enter(e):
            username.delete(0, "end")

        def when_leave(e):
            name = username.get()
            if name == "":
                username.insert(0, "Username")

        username.bind("<FocusIn>", when_enter)
        username.bind("<FocusOut>", when_leave)

        # Underline Username
        tk.Frame(self.register_frame, width=295, height=2, bg="black").place(x=25, y=107)

        # Input Password
        password = tk.Entry(self.register_frame, width=25, fg="black", border=0, bg="white", font=("Jokerman", 11))
        password.place(x=30, y=150)
        password.insert(0, "Password")

        # Interact Password When Click
        def when_enter(e):
            password.delete(0, "end")
            password.config(show="*")

        def when_leave(e):
            passcode = password.get()
            if passcode == "":
                password.insert(0, "Password")
                password.config(show="")

        password.bind("<FocusIn>", when_enter)
        password.bind("<FocusOut>", when_leave)

        # Underline Password
        tk.Frame(self.register_frame, width=295, height=2, bg="black").place(x=25, y=177)

        # Input Confirm Password
        con_password = tk.Entry(self.register_frame, width=25, fg="black", border=0, bg="white", font=("Jokerman", 11))
        con_password.place(x=30, y=220)
        con_password.insert(0, "Confirm Password")

        # Interact Confirm Password When Click
        def when_enter(e):
            con_password.delete(0, "end")
            con_password.config(show="*")

        def when_leave(e):
            con_passcode = con_password.get()
            if con_passcode == "":
                con_password.insert(0, "Confirm Password")
                con_password.config(show="")

        con_password.bind("<FocusIn>", when_enter)
        con_password.bind("<FocusOut>", when_leave)

        # Underline Password
        tk.Frame(self.register_frame, width=295, height=2, bg="black").place(x=25, y=247)

        # Sign up
        signup = tk.Button(self.register_frame, width=39, pady=7, text="Sign Up", command=lambda: self.register(username, password, con_password), bg="#57a1f8", fg="white", border=0)
        signup.place(x=35, y=274)

        # Label "I Have An Account"
        no_account = tk.Label(self.register_frame, text="I Have An Account", bg="white", fg="black", font=("Centaur", 10))
        no_account.place(x=175, y=320)

        # Sign In
        signin = tk.Button(self.register_frame, width=6, text="Sign In", command=self.show_login_screen, bg="White", fg="#57a1f8", cursor="hand2", border=0)
        signin.place(x=270, y=320)

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def login(self, username_entry, password_entry):
        username = username_entry.get()
        password = password_entry.get()

        self.success = False
        try:
            with open("logindata.txt", "r") as file:
                for line in file:
                    a, b = line.strip().split(",")
                    if a == username and b == password:
                        self.success = True
                        break
        except FileNotFoundError:
            messagebox.showerror(title="Error", message="User data file not found.")
            return

        if self.success:
            self.clear_window()
            self.root.destroy()
            MajorSelection()
        else:
            messagebox.showerror(title="Invalid", message="Invalid username/password")

    def register(self, username_entry, password_entry, con_password_entry):
        username = username_entry.get()
        password = password_entry.get()
        con_password = con_password_entry.get()

        if not all([username, password, con_password]):
            messagebox.showerror(title="Invalid", message="Please insert username and password")
        elif password == con_password:
            with open("logindata.txt", "a") as file:
                file.write(f"{username},{password}\n")
            messagebox.showinfo(title="Registered", message="Successfully registered")
            self.show_login_screen()
        else:
            messagebox.showerror(title="Invalid", message="Password and confirm password do not match")


class MajorSelection:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Pemilihan Jurusan Universitas Sigma")
        self.window.geometry("1440x800")
        self.window.configure(bg="white")
        self.window.resizable(False,False)

        # Load Workbook
        try:
            self.wb = load_workbook("data.xlsx")
            self.ws = self.wb.active
        except FileNotFoundError:
            messagebox.showerror("Error", "Data file not found.")
            self.window.destroy()
            return

        # Print In Excel (Debug)
        for row in self.ws.iter_rows(values_only=True):
            print(row)

        self.create_widgets()
        self.window.mainloop()

    def create_widgets(self):
        # Labels
        tk.Label(bg="#FFFFFF", width=700).place(x=0, y=820)
        tk.Label(bg="#FFFFFF", width=700).place(x=0, y=0)
        tk.Label(bg="#FFFFFF", width=10, height=700).place(x=0, y=0)
        tk.Label(bg="#FFFFFF", width=10, height=700).place(x=1460, y=0)
        tk.Label(bg="#FFFFFF", text="Pemilihan Jurusan Universitas Sigma",fg="#57a1f8",font=("Poetsen One", 20)).place(x=530, y=50)
        
        # Logo Image Big Brain
        img_path = r"C:\Users\hatab\OneDrive\Pictures\Saved Pictures\Cumlawd.jpg"
        if os.path.exists(img_path):
            imgBigBrain = Image.open(img_path)
            self.imgBigBrain = ImageTk.PhotoImage(imgBigBrain)
            tk.Label(self.window, image=self.imgBigBrain, bg="white").place(x=300, y=200)
        else:
            messagebox.showwarning("Warning", "Logo image not found. Image will not be displayed.")
        
        # Logo Image Sigma Major
        img_path = r"C:\Users\hatab\OneDrive\Pictures\Saved Pictures\Sigma-sigma.jpg"
        if os.path.exists(img_path):
            imgSigmaMajor = Image.open(img_path)
            self.imgSigmaMajor = ImageTk.PhotoImage(imgSigmaMajor)
            tk.Label(self.window, image=self.imgSigmaMajor, bg="white").place(x=220, y=500)
        else:
            messagebox.showwarning("Warning", "Logo image not found. Image will not be displayed.")
            
        self.major_frame = tk.Frame(self.window, bg = "white",width = 700, height = 600)
        self.major_frame.place(x=700,y=100)
        
        tk.Label(self.major_frame, bg="#FFFFFF", text="Nama : ",fg="#57a1f8", font=("Poetsen One", 15)).place(x=30, y=80)
        tk.Label(self.major_frame, bg="#FFFFFF", text="NIK : ",fg="#57a1f8", font=("Poetsen One", 15)).place(x=30, y=180)
        tk.Label(self.major_frame, bg="#FFFFFF", text="No Telp : ",fg="#57a1f8", font=("Poetsen One", 15)).place(x=30, y=280)
        tk.Label(self.major_frame, bg="#FFFFFF", text="Jurusan : ",fg="#57a1f8", font=("Poetsen One", 15)).place(x=30, y=380)

        # Input Nama
        self.name = tk.StringVar()
        tk.Entry(self.major_frame, width=40, textvariable=self.name,border=0, font=("Poetsen One", 15)).place(x=150, y=80)
        
        # Underline Nama
        tk.Frame(self.major_frame, width=455, height=2, bg="black").place(x=145, y=107)

        # Input NIK
        self.NoIK = tk.StringVar()
        tk.Entry(self.major_frame, width=40, textvariable=self.NoIK,border=0, font=("Poetsen One", 15)).place(x=150, y=180)
        
        # Underline NIK
        tk.Frame(self.major_frame, width=455, height=2, bg="black").place(x=145, y=207)

        # Input No Telp
        self.NoTelp = tk.StringVar()
        tk.Entry(self.major_frame,width=40, textvariable=self.NoTelp,border=0, font=("Poetsen One", 15)).place(x=150, y=280)
        
        # Underline No Telp
        tk.Frame(self.major_frame, width=455, height=2, bg="black").place(x=145, y=307)

        # Option Jurusan
        self.Jurusan = tk.StringVar()
        OptionJurusan = ttk.Combobox(self.major_frame, width=39, font=("Poetsen One", 15), textvariable=self.Jurusan, state="readonly")
        OptionJurusan.place(x=150, y=380)
        OptionJurusan['values'] = ('Teknik Komputer', 'Teknik Mesin', 'Teknik Kimia', 'Teknik Industri', 'Teknik Sipil', 'Teknik Kapal', 'Teknik Lingkungan', 'PWK', 'Kedokteran', 'Hukum', 'Hubungan Internasional', 'Management', 'Business')

        # Tombol Pendaftaran
        tk.Button(self.major_frame, width=20, pady=7, bg="#EB984E", command=self.Confirm_Major, text="Submit").place(x=160, y=480)

        # Tombol Keluar
        tk.Button(self.major_frame, width=20, pady=7, bg="Red", command=self.window.destroy, text="Exit").place(x=400, y=480)

    def Confirm_Major(self):
        if not all([self.name.get(), self.NoIK.get(), self.NoTelp.get()]):
            messagebox.showerror(title="Data masih kosong", message="HARAP ISI DATA ANDA UNTUK MELANJUTKAN PEMILIHAN JURUSAN")
        elif not self.Jurusan.get():
            messagebox.showerror(title="PILIH JURUSAN!", message="HARAP PILIH JURUSAN UNTUK MELANJUTKAN PEMILIHAN JURUSAN")
        else:
            # Write Excel
            self.ws.append([self.name.get(), self.NoIK.get(), self.NoTelp.get(), self.Jurusan.get()])
            self.wb.save("data.xlsx")  # Save the Excel file
            messagebox.showinfo(title=f"Selamat {self.name.get()} anda sudah terdaftar", message=f"Anda telah memilih jurusan {self.Jurusan.get()}")
            
            # Close Window After Done
            self.window.destroy()

Login_System()
